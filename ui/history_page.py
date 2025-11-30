from PyQt6.QtWidgets import (
    QWidget, 
    QVBoxLayout, 
    QHBoxLayout,
    QLabel, 
    QTableWidget, 
    QTableWidgetItem, 
    QPushButton, 
    QHeaderView,
    QFileDialog,
    QMessageBox,
    QApplication,
    QStyle,
    QComboBox,
    QLineEdit,
    QGroupBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtCore import Qt
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from database.database import create_connection, get_record_by_id, delete_record_by_id, get_paginated_records, get_record_count, get_distinct_result_types, get_all_records_filtered

class HistoryPage(QWidget):
    def __init__(self):
        super().__init__()

        self.records_per_page = 15
        self.current_page = 1
        self.total_pages = 1
        self.initial_load_done = False
        
        self._init_ui()
        # Initial population is now handled by showEvent to prevent sizing bugs
        
    def showEvent(self, event):
        """Override showEvent to populate the table on first show."""
        if not self.initial_load_done:
            self._populate_filter_combo() # Populate filters on first show
            self.apply_filters_and_search()
            self.initial_load_done = True
        super().showEvent(event)

    def _init_ui(self):
        layout = QVBoxLayout(self)
        
        top_bar_layout = QHBoxLayout()
        title = QLabel("Detection History")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #1E3A8A;")
        
        top_bar_layout.addWidget(title, alignment=Qt.AlignmentFlag.AlignLeft)
        top_bar_layout.addStretch()
        layout.addLayout(top_bar_layout)
        
        controls_group = QGroupBox("Filter, Search & Actions")
        controls_layout = QHBoxLayout(controls_group)
        
        self.filter_combo = QComboBox()
        self.filter_combo.currentTextChanged.connect(self.apply_filters_and_search)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by ID, date, result...")
        self.search_input.textChanged.connect(self.apply_filters_and_search)
        
        # Tombol Export Excel Baru
        self.export_all_btn = QPushButton("Export All to Excel")
        self.export_all_btn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        self.export_all_btn.setStyleSheet("background-color: #10B981; color: white; font-weight: bold; padding: 5px 10px; border-radius: 4px;")
        self.export_all_btn.clicked.connect(self.export_all_to_excel)

        controls_layout.addWidget(QLabel("Filter by Result:"))
        controls_layout.addWidget(self.filter_combo)
        controls_layout.addSpacing(20)
        controls_layout.addWidget(QLabel("Search:"))
        controls_layout.addWidget(self.search_input, 1)
        controls_layout.addSpacing(20)
        controls_layout.addWidget(self.export_all_btn)
        
        layout.addWidget(controls_group)
        layout.addSpacing(10)

        self.history_table = QTableWidget()
        self.history_table.setColumnCount(4)
        self.history_table.setHorizontalHeaderLabels(["No", "Date and Time", "Result", "Action"])
        self.history_table.setStyleSheet("""
            QTableWidget { font-size: 14px; border: 1px solid #D1D5DB; border-radius: 8px; gridline-color: #E5E7EB; }
            QHeaderView::section { background-color: #F3F4F6; padding: 10px; border: none; font-weight: bold; font-size: 14px; color: #374151; }
            QTableWidget::item { padding: 10px; color: #111827; }
            QTableWidget::alternating-row-color { background-color: #F9FAFB; }
        """)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setShowGrid(False)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.verticalHeader().setDefaultSectionSize(50) # Increased default row height to 50
        
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.history_table.setColumnWidth(3, 120)

        layout.addWidget(self.history_table)

        pagination_layout = QHBoxLayout()
        self.prev_button = QPushButton("← Previous")
        self.prev_button.clicked.connect(self.prev_page)
        self.page_label = QLabel("Page 1 of 1")
        self.next_button = QPushButton("Next →")
        self.next_button.clicked.connect(self.next_page)

        pagination_layout.addStretch()
        pagination_layout.addWidget(self.prev_button)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.next_button)
        pagination_layout.addStretch()

        layout.addLayout(pagination_layout)

    def _populate_filter_combo(self):
        conn = create_connection()
        if conn:
            try:
                # Block signals to prevent triggering a refresh while populating
                self.filter_combo.blockSignals(True)
                
                current_selection = self.filter_combo.currentText()
                results = get_distinct_result_types(conn)
                
                self.filter_combo.clear()
                self.filter_combo.addItem("All")
                self.filter_combo.addItems(results)
                
                # Restore previous selection if it still exists
                index = self.filter_combo.findText(current_selection)
                if index != -1:
                    self.filter_combo.setCurrentIndex(index)

                self.filter_combo.blockSignals(False)
            finally:
                conn.close()

    def apply_filters_and_search(self):
        self.current_page = 1
        self.populate_table()

    def populate_table(self):
        filter_text = self.filter_combo.currentText()
        search_query = self.search_input.text()

        conn = create_connection()
        if not conn:
            return

        try:
            total_records = get_record_count(conn, filter_text, search_query)
            self.total_pages = math.ceil(total_records / self.records_per_page) if total_records > 0 else 1
            
            if self.current_page > self.total_pages: self.current_page = self.total_pages
            if self.current_page < 1: self.current_page = 1

            offset = (self.current_page - 1) * self.records_per_page
            records = get_paginated_records(conn, offset, self.records_per_page, filter_text, search_query)
            
            self.history_table.setRowCount(0)
            
            if not records:
                # Tampilkan pesan kosong di baris pertama jika mau, atau biarkan kosong
                return

            self.history_table.setRowCount(len(records))
            for i, record in enumerate(records):
                record_id = record[0]
                timestamp = record[1]
                result = record[2]
                
                # ID
                self.history_table.setItem(i, 0, QTableWidgetItem(str(record_id)))
                
                # Waktu
                self.history_table.setItem(i, 1, QTableWidgetItem(str(timestamp)))
                
                # Hasil
                item_res = QTableWidgetItem(str(result))
                if "Terdeteksi" in str(result):
                    item_res.setForeground(Qt.GlobalColor.red)
                else:
                    item_res.setForeground(Qt.GlobalColor.darkGreen)
                
                font = QFont()
                font.setBold(True)
                item_res.setFont(font)
                
                self.history_table.setItem(i, 2, item_res)
                
                # Aksi
                self._create_action_buttons(i, record_id)
                
                # Align Center
                for c in range(3):
                    item = self.history_table.item(i, c)
                    if item: item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            self.update_pagination_ui()
            
        except Exception as e:
            print(f"Error loading history: {e}")
        finally:
            conn.close()

    def _create_action_buttons(self, row, record_id):
        actions_widget = QWidget()
        layout = QHBoxLayout(actions_widget)
        layout.setContentsMargins(5, 2, 5, 2)
        layout.setSpacing(10)
        
        # Tombol Detail
        btn_detail = QPushButton("📄 Detail")
        btn_detail.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_detail.setStyleSheet("background-color: #3B82F6; color: white; border-radius: 4px; padding: 4px;")
        btn_detail.clicked.connect(lambda: self.show_record_detail(record_id))
        
        # Tombol Hapus
        btn_delete = QPushButton("🗑️")
        btn_delete.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_delete.setFixedWidth(30)
        btn_delete.setStyleSheet("background-color: #EF4444; color: white; border-radius: 4px; padding: 4px;")
        btn_delete.clicked.connect(lambda: self.delete_record_confirmation(record_id))
        
        layout.addWidget(btn_detail)
        layout.addWidget(btn_delete)
        layout.addStretch()
        
        self.history_table.setCellWidget(row, 3, actions_widget)

    def _get_bold_font(self):
        font = self.history_table.font()
        font.setBold(True)
        return font

    def update_pagination_ui(self):
        if self.total_pages <= 1:
            self.page_label.setText("Page 1 of 1")
            self.prev_button.setEnabled(False)
            self.next_button.setEnabled(False)
        else:
            self.page_label.setText(f"Page {self.current_page} of {self.total_pages}")
            self.prev_button.setEnabled(self.current_page > 1)
            self.next_button.setEnabled(self.current_page < self.total_pages)

    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.populate_table()

    def next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.populate_table()
            
    def _set_icon_button_style(self, button, tooltip):
        button.setToolTip(tooltip)
        button.setCursor(Qt.CursorShape.PointingHandCursor)
        button.setStyleSheet("""
            QPushButton { background-color: transparent; border: none; padding: 4px; border-radius: 15px; }
            QPushButton:hover { background-color: #E5E7EB; }
        """)
        button.setFixedSize(30, 30)
        
    def show_record_detail(self, record_id):
        try:
            detail_dialog = DetailPage(record_id, self)
            detail_dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Tidak dapat membuka halaman detail: {e}")

    def delete_record_confirmation(self, record_id):
        reply = QMessageBox.question(self, 'Konfirmasi Hapus',
                                     f"Anda yakin ingin menghapus record dengan ID {record_id}?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.delete_record(record_id)

    def delete_record(self, record_id):
        conn = create_connection()
        if not conn:
            QMessageBox.warning(self, "Error Database", "Tidak dapat terhubung ke database untuk menghapus.")
            return

        try:
            rows_affected = delete_record_by_id(conn, record_id)
            conn.close()
            if rows_affected > 0:
                self.apply_filters_and_search()
            else:
                QMessageBox.warning(self, "Hapus Gagal", f"Record dengan ID {record_id} tidak ditemukan atau gagal dihapus.")
        except Exception as e:
            QMessageBox.critical(self, "Error Hapus", f"Terjadi kesalahan saat menghapus data: {e}")

    def export_all_to_excel(self):
        """Eksport semua data (sesuai filter saat ini) ke file Excel (.xlsx)."""
        filter_text = self.filter_combo.currentText()
        search_query = self.search_input.text()
        
        conn = create_connection()
        if not conn:
            return
            
        # Ambil semua data yang sesuai filter (bukan cuma 1 halaman)
        records = get_all_records_filtered(conn, filter_text, search_query)
        conn.close()
        
        if not records:
            QMessageBox.information(self, "Info", "Tidak ada data untuk diekspor.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "e_nose_history.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detection History"
            
            # Header Style
            headers = ["ID", "Waktu Deteksi", "Hasil Prediksi", "Raw Data (Preview)"]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Isi Data
            for row_idx, record in enumerate(records, start=2):
                ws.cell(row=row_idx, column=1, value=record[0]) # ID
                ws.cell(row=row_idx, column=2, value=record[1]) # Timestamp
                ws.cell(row=row_idx, column=3, value=record[2]) # Result
                
                # Raw data kadang panjang, kita potong dikit buat preview
                raw_data_preview = record[3][:100] + "..." if len(record[3]) > 100 else record[3]
                ws.cell(row=row_idx, column=4, value=raw_data_preview)

            # Auto adjust column width (biar rapi)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50) # Max width 50 biar gak lebar banget

            wb.save(path)
            QMessageBox.information(self, "Sukses", f"Data berhasil diekspor ke:\n{path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error Export", f"Gagal menyimpan file Excel:\n{e}")

    def export_single_record(self, record_id):
        conn = create_connection()
        if not conn:
            QMessageBox.warning(self, "Error Database", "Tidak dapat terhubung ke database untuk mengekspor.")
            return

        record = get_record_by_id(conn, record_id)
        conn.close()

        if not record:
            QMessageBox.warning(self, "Record Tidak Ditemukan", f"Record dengan ID {record_id} tidak ditemukan untuk diekspor.")
            return
            
        path, _ = QFileDialog.getSaveFileName(self, f"Simpan Record {record_id} ke CSV", f"record_{record_id}.csv", "CSV Files (*.csv)")

        if path:
            try:
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['ID', 'Timestamp', 'Result', 'RawData'])
                    writer.writerow(record)
                QMessageBox.information(self, "Ekspor Berhasil", f"Record {record_id} berhasil diekspor ke {path}")
            except Exception as e:
                QMessageBox.critical(self, "Error Ekspor", f"Terjadi kesalahan saat mengekspor record: {e}")